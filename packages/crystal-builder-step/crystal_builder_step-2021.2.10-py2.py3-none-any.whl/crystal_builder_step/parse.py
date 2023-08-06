#!/usr/bin/env python3

import json
import string
import textwrap

import CifFile
import openpyxl

subscript = [
    '\N{Subscript Zero}',
    '\N{Subscript One}',
    '\N{Subscript Two}',
    '\N{Subscript Three}',
    '\N{Subscript Four}',
    '\N{Subscript Five}',
    '\N{Subscript Six}',
    '\N{Subscript Seven}',
    '\N{Subscript Eight}',
    '\N{Subscript Nine}',
]

bib = {}
bib['A3B_hP8_158_d_a'] = """\
@Article{A3B_hP8_158_d_a,
    author ="Fletcher, J. M. and Gardner, W. E. and Fox, A. C. and Topping, G.",
    title  ="X-Ray{,} infrared{,} and magnetic studies of α- and β-ruthenium trichloride",
    journal  ="J. Chem. Soc. A",
    year  ="1967",
    issue  ="0",
    pages  ="1038-1045",
    publisher  ="The Royal Society of Chemistry",
    doi  ="10.1039/J19670001038",
    url  ="http://dx.doi.org/10.1039/J19670001038"
}
"""  # noqa: E501

bib['A4B_oI20_74_beh_e'] = """\
@inbook{A4B_oI20_74_beh_e,
    place={Berlin &amp; Heidelberg, Germany},
    edition={8th},
    series={Gmelin Handbook of Inorganic and Organometallic Chemistry},
    title={Supplement Volume B2},
    volume={B2},
    booktitle={Alloys of Uranium with Alkali Metals, Alkaline Earths, and Elements of Main Groups III and IV},
    publisher={Springer},
    author={Borgstedt, Hans Ulrich and Wedemeyer, Horst and Buschbeck, Karl-Christian},
    year={1989},
    pages={1–333},
    collection={Gmelin Handbook of Inorganic and Organometallic Chemistry}
}

"""  # noqa: E501

bib['A4B_tP10_125_m_a'] = """\
@article{A4B_tP10_125_m_a,
    author = {R. Graham and G. C. S. Waghorn and P. T. Davies},
    title = {An X–ray investigation of the lead–platinum system},
    journal = {Acta Crystallographica},
    year = 1954,
    volume = {7},
    number = {10},
    pages = {634--635},
    month = {Oct},
    doi = {10.1107/S0365110X54002137},
    url = {https://doi.org/10.1107/S0365110X54002137},
}

"""

bib['A2B_oP12_62_2c_c.SrH2'] = """\
@inbook{A2B_oP12_62_2c_c.SrH2,
    author={R. C. Ropp},
    title={Encyclopedia of the Alkaline Earth Compounds},
    place={Oxford},
    publisher={Elsevier},
    year=2013,
    pages={30--31}
}

"""

bib['AB8C2_oC22_35_a_ab3e_e'] = """\
@article{AB8C2_oC22_35_a_ab3e_e,
    title = {CHEMICAL AND STRUCTURAL INVESTIGATION OF THE AB$sub 2$O$sub 8$
             COMPOUNDS OF Mo-V-O, U-V-O, AND U-Mo-O SYSTEMS.},
    author = {Mahe-Pailleret, P},
    abstractNote = {},
    doi = {},
    journal = {Rev. Chim. Miner. 7: 807-48(1970).},
    number = ,
    volume = ,
    place = {France},
    year = {1970},
    month = {1}
}

"""

bib['AB_hP6_144_a_a'] = """\
@article{doi:10.1063/1.46096,
    author = {Kusaba,Keiji  and Weidner,Donald J. },
    title = {Structure of high pressure phase I in ZnTe},
    journal = {AIP Conference Proceedings},
    volume = {309},
    number = {1},
    pages = {553-556},
    year = {1994},
    doi = {10.1063/1.46096},
    URL = {https://aip.scitation.org/doi/abs/10.1063/1.46096},
    eprint = {https://aip.scitation.org/doi/pdf/10.1063/1.46096}
}

"""

bib['A12B36CD12_cF488_210_h_3h_a_fg'] = """\
@article{A12B36CD12_cF488_210_h_3h_a_fg,
    author = {Tiritiris, Ioannis and Schleid, Thomas},
    title = {Synthesis, Crystal Structure, and Thermal Decomposition of Mg(H2O)6 [B12H12]×6H2O.},
    journal = {ChemInform},
    volume = {35},
    number = {25},
    pages = {},
    doi = {10.1002/chin.200425008},
    url = {https://onlinelibrary.wiley.com/doi/abs/10.1002/chin.200425008},
    eprint = {https://onlinelibrary.wiley.com/doi/pdf/10.1002/chin.200425008},
    abstract = {Abstract For Abstract see ChemInform Abstract in Full Text.},
    year = {2004}
}

"""  # noqa: E501

bib['AB3_tP32_86_g_3g'] = """\
@article{AB3_tP32_86_g_3g,
  title={PHASE DIAGRAM OF THE TITANIUM-PHOSPHORUS SYSTEM},
  author={EREMENKO, VN and LISTOVNICHII, VE},
  journal={DOPOVIDI AKAD NAUK UKR SSR},
  number={9},
  pages={1176--1179},
  year={1965}
}

"""

bib['AB32CD4E8_tP184_93_i_16p_af_2p_4p'] = """\
@article{AB32CD4E8_tP184_93_i_16p_af_2p_4p,
  title={Crystal and molecular structures of [AsPh 4][Ln (S 2 PMe 2) 4](Ln= Ce or Tm) and their comparison with results obtained from paramagnetic nuclear magnetic resonance data in solution},
  author={Spiliadis, Stavros and Pinkerton, A Alan and Schwarzenbach, Dieter},
  journal={Journal of the Chemical Society, Dalton Transactions},
  number={9},
  pages={1809--1813},
  year={1982},
  publisher={Royal Society of Chemistry}
}

"""  # noqa: E501

bib['A3B_oC64_66_gi2lm_2l'] = """\
@article{A3B_oC64_66_gi2lm_2l,
  title={Pressure-induced metallization of dense (H 2 S) 2 H 2 with high-Tc  superconductivity},
  author={Duan, Defang and Liu, Yunxian and Tian, Fubo and Li, Da and Huang, Xiaoli and Zhao, Zhonglong and Yu, Hongyu and Liu, Bingbing and Tian, Wenjing and Cui, Tian},
  journal={Scientific reports},
  volume={4},
  pages={6968},
  year={2014},
  publisher={Nature Publishing Group}
}

"""  # noqa: E501

bib['ABC2_hP4_164_a_b_d'] = """\
@phdthesis{ABC2_hP4_164_a_b_d,
  title={Intermetallic compounds by reductive annealing},
  author={Kift, Rebecca Louise},
  year={2010},
  school={University of Hull}
}

"""


def formula_subscripts(formula):
    result = ''
    for char in formula:
        if char in '0123456789':
            result += subscript[int(char)]
        else:
            result += char
    return result


def run(filename='prototypes.xlsx'):
    data = {}

    wb = openpyxl.load_workbook(filename=filename)
    ws = wb['Sheet1']

    row_max = ws.max_row

    data = {
        'prototype': [],
        'nSpecies': [],
        'nAtoms': [],
        'Pearson symbol': [],
        'Strukturbericht designation': [],
        'AFLOW prototype': [],
        'space group symbol': [],
        'space group number': [],
        'notes': []
    }
    hyperlink = []

    for row in ws.iter_rows(min_row=2, max_row=row_max):
        hyperlink.append(row[0].hyperlink.display)
        for cell, column in zip(row, data.keys()):
            data[column].append(cell.value)

    data['hyperlink'] = hyperlink

    # Now clean up the data
    tmp = []
    # Characters in notes
    for value in data['notes']:
        tmp.append(value.replace('\u00a0', ' '))
    data['notes'] = tmp

    # Subscripts in formulas of prototypes
    tmp = []
    for value in data['prototype']:
        value = formula_subscripts(value)
        print(value)
        tmp.append(value)
    data['prototype'] = tmp

    # doubling of Strukturbericht designations
    tmp = []
    for value in data['Strukturbericht designation']:
        if value is not None:
            length = int(len(value) / 2)
            value = value[0:length]
        tmp.append(value)
    data['Strukturbericht designation'] = tmp

    # doubling of space group symbols
    tmp = []
    for value in data['space group symbol']:
        if value is not None:
            length = int(len(value) / 2)
            value = value[0:length].replace('\u00af', '\N{Combining Overline}')
        print(value)
        tmp.append(value)
    data['space group symbol'] = tmp

    with open('prototypes.json', 'w') as fd:
        json.dump(data, fd, indent=4)


def download():
    """Download the CIF files from AFLOW"""
    # Get the data from the json
    with open('prototypes.json', 'r') as fd:
        data = json.load(fd)

    # Fetch the cif file
    for hyperlink in data['hyperlink']:
        url = urllib.parse.urlsplit(hyperlink)
        path = url.path
        root, extension = os.path.splitext(os.path.basename(path))
        filename = root + '.cif'
        new_path = os.path.join('prototypes', filename)
        if os.path.exists(new_path):
            continue
        new_url = urllib.parse.urljoin(hyperlink, 'CIF/' + filename)
        print(f'{new_url} --> {new_path}')
        try:
            with urllib.request.urlopen(new_url) as response:
                with open(new_path, 'w') as fd:
                    fd.write(response.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            print(f'#### Error: {e}')


def get_bibtex(prototype):
    """Create the BibTex for the prototype

    Example:

    @article{doi:10.1139/v68-576,
        author = {Knop, Osvald and Reid, K. I. G. and Sutarno and Nakagawa,
                 Yasuaki},
        title = {Chalkogenides of the transition elements. VI. X-Ray, neutron,
                 and magnetic investigation of the spinels Co3O4, NiCo2O4,
                 Co3S4, and NiCo2S4},
        journal = {Canadian Journal of Chemistry},
        volume = {46},
        number = {22},
        pages = {3463-3476},
        year = {1968},
    }
    """
    try:
        cif = CifFile.ReadCif(prototype + '.cif')
    except Exception:
        print(f'Error in {prototype}.cif\n\n')
        raise
    tmp = cif['findsym-output']
    authors = tmp['_publ_author_name']
    if '_journal_name_full' in tmp:
        journal = tmp['_journal_name_full']
    else:
        journal = tmp['_journal_name_full_name']
    if '_journal_volume' in tmp:
        volume = tmp['_journal_volume']
    else:
        volume = 'n/a'
        print(f'{prototype} has no journal volume')
    year = tmp['_journal_year']
    first = tmp['_journal_page_first']
    last = tmp['_journal_page_last']
    title = tmp['_publ_Section_title']

    author_list = ' and '.join(authors)
    pages = f'{first}-{last}'

    template = string.Template(
        """\
        @article{$prototype,
        author = {$author_list},
        title = {$title},
        journal = {$journal},
        volume = {$volume},
        pages = {$pages},
        year = $year,
        }
        """
    )

    citation = template.substitute(
        prototype=prototype,
        author_list=author_list,
        title=title.strip(),
        journal=journal.strip(),
        volume=volume,
        pages=pages,
        year=year
    )

    return citation


def test():
    import bibtexparser
    import reference_handler

    with open('references.bib') as fd:
        data = fd.read()

    tmp = bibtexparser.loads(data).entries_dict
    writer = bibtexparser.bwriter.BibTexWriter()
    biblio = {}
    for key, data in tmp.items():
        biblio[key] = writer._entry_to_bibtex(data)

    references = reference_handler.Reference_Handler(':memory:')
    for key, citation in biblio.items():
        references.cite(
            raw=citation,
            alias=key,
            module='parser',
            level=1,
            note=f"The principle citation for {key}."
        )

    citations = references.dump(fmt='text')
    for citation, text, count, level in citations:
        print(f"{citation} {textwrap.fill(text)}\n")


if __name__ == "__main__":  # pragma: no cover
    import os.path
    import urllib.error
    import urllib.parse
    import urllib.request

    test()

    # if True:
    #     test()
    # elif False:
    #     bibtex = ''
    #     with open('prototypes.json', 'r') as fd:
    #         data = json.load(fd)

    #     for prototype in data['AFLOW prototype']:
    #         print(prototype)
    #         if prototype in bib:
    #             text = bib[prototype]
    #         else:
    #             text = get_bibtex(prototype)
    #         bibtex += text
    #     with open('references2.bib', 'w') as fd:
    #         fd.write(bibtex)
    # elif False:
    #     run()
    # elif False:
    #     # Fix the prototypes to have added formula part in some cases
    #     # Get the data from the json
    #     with open('prototypes.json', 'r') as fd:
    #         data = json.load(fd)

    #     aflow_prototypes = []
    #     for hyperlink in data['hyperlink']:
    #         url = urllib.parse.urlsplit(hyperlink)
    #         path = url.path
    #         root, extension = os.path.splitext(os.path.basename(path))
    #         aflow_prototypes.append(root)

    #     data['AFLOW prototype'] = aflow_prototypes

    #     with open('prototypes2.json', 'w') as fd:
    #         json.dump(data, fd, indent=4)
    # else:
    #     # Get the data from the json
    #     with open('prototypes2.json', 'r') as fd:
    #         data = json.load(fd)

    #     cp = data['cell parameters'] = []
    #     sites = data['sites'] = []
    #     # Fetch the cif file to get the adjustable parameters
    #     for hyperlink in data['hyperlink']:
    #         url = urllib.parse.urlsplit(hyperlink)
    #         path = url.path
    #         root, extension = os.path.splitext(os.path.basename(path))
    #         filename = root + '.cif'
    #         new_path = os.path.join('.', filename)
    #         if not os.path.exists(new_path):
    #             print(f'{new_path} does not exist')
    #         else:
    #             try:
    #                 cif = CifFile.ReadCif(new_path)
    #             except Exception:
    #                 print(f'Error in {new_path}\n\n')
    #                 raise
    #             tmp = cif['findsym-output']
    #             params = tmp['_aflow_params'].split(',')
    #             values = tmp['_aflow_params_values'].split(',')

    #             cell_parameters = []
    #             for param, value in zip(params, values):
    #                 if param[0] in ('x', 'y', 'z'):
    #                     break
    #                 if param[0] == '\\':
    #                     cell_parameters.append((param[1:], value))
    #                 elif param == 'b/a':
    #                     cell_parameters.append(('b', tmp['_cell_length_b']))
    #                 elif param == 'c/a':
    #                     cell_parameters.append(('c', tmp['_cell_length_c']))
    #                 else:
    #                     cell_parameters.append((param, value))
    #             cp.append(cell_parameters)

    #             site_data = []
    #             for symbol, site, mult in zip(
    #                 tmp['_atom_site_type_symbol'],
    #                 tmp['_atom_site_Wyckoff_label'],
    #                 tmp['_atom_site_symmetry_multiplicity']
    #             ):
    #                 site_data.append((site, int(mult), symbol))
    #             sites.append(site_data)

    #     with open('prototypes3.json', 'w') as fd:
    #         json.dump(data, fd, indent=4)
