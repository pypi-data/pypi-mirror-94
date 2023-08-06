from openpyxl import load_workbook  # type: ignore
from typing import List


class Worksheet:
    def __init__(self) -> None:
        self.sheet_items: List[dict] = []

    def __repr__(self) -> str:
        return str(self.sheet_items)

    def xlsx_to_dict(self, path):
        """
        Read a Worksheet and return it as array of dictionaries
        :param self: Worksheet Object
        :param path: Path to XLSX file
        :return: Array of rows as dictionaries
        """
        book = load_workbook(path)
        sheet = book.active
        rows = sheet.max_row
        cols = sheet.max_column

        self.sheet_items = []

        def item(row, col):
            return (
                sheet.cell(row=1, column=col).value,
                str(sheet.cell(row=row, column=col).value)
            )

        for row in range(2, rows + 1):
            self.sheet_items.append(
                dict(item(row, col) for col in range(1, cols + 1))
            )

        return self

    def csv_to_dict(self, csv_file, delimiter=','):
        """
        Read a CSV and return it as array of dictionaries
        :param self: CSV Object
        :param csv_file: CSV file
        :return: Array of rows as dictionaries
        """
        import csv
        self.sheet_items = []

        dict_reader = csv.DictReader(csv_file, delimiter=delimiter)
        self.sheet_items = list(dict_reader)
        return self.sheet_items

    def headers(self):
        return self.sheet_items[0]
